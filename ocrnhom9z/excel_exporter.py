# excel_exporter.py - Xuáº¥t dá»¯ liá»‡u ra Excel

import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import *

class ExcelExporter:
    """Class xá»­ lÃ½ xuáº¥t dá»¯ liá»‡u ra Excel"""
    
    def __init__(self):
        pass
    
    def export_to_excel(self, df, file_path):
        """Xuáº¥t DataFrame ra file Excel vá»›i Ä‘á»‹nh dáº¡ng Ä‘áº¹p"""
        if df.empty:
            raise ValueError("KhÃ´ng cÃ³ dá»¯ liá»‡u Ä‘á»ƒ xuáº¥t")
        
        try:
            # Táº¡o tÃªn sheet vá»›i timestamp
            sheet_name = f"{EXCEL_SHEET_NAME_PREFIX}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Loáº¡i bá» cÃ¡c cá»™t khÃ´ng cÃ³ dá»¯ liá»‡u
            df_export = self._remove_empty_columns(df)

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Xuáº¥t dá»¯ liá»‡u
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)

                # Láº¥y worksheet Ä‘á»ƒ Ä‘á»‹nh dáº¡ng
                worksheet = writer.sheets[sheet_name]

                # Ãp dá»¥ng Ä‘á»‹nh dáº¡ng
                self._format_worksheet(worksheet, df_export)
                
            return True, f"ÄÃ£ xuáº¥t thÃ nh cÃ´ng {len(df)} sinh viÃªn"
            
        except Exception as e:
            return False, f"Lá»—i xuáº¥t Excel: {str(e)}"

    def _remove_empty_columns(self, df):
        """Loáº¡i bá» cÃ¡c cá»™t khÃ´ng cÃ³ dá»¯ liá»‡u"""
        df_clean = df.copy()

        # Danh sÃ¡ch cÃ¡c cá»™t cÃ³ thá»ƒ bá»‹ trá»‘ng
        potential_empty_cols = ['KT2', 'KDT', 'KT3', 'KT4', 'TH1', 'TH2', 'TH3', 'TH4']

        columns_to_remove = []

        for col in potential_empty_cols:
            if col in df_clean.columns:
                # Convert sang string vÃ  kiá»ƒm tra
                col_str = df_clean[col].fillna('').astype(str)

                # Kiá»ƒm tra cÃ¡c giÃ¡ trá»‹ cÃ³ thá»ƒ coi lÃ  "trá»‘ng"
                empty_indicators = ['', 'nan', 'None', '0.0', '1.0', 'NaN']

                # Äáº¿m sá»‘ giÃ¡ trá»‹ thá»±c sá»± cÃ³ nghÄ©a
                meaningful_values = []
                for val in col_str:
                    val_clean = str(val).strip()
                    if val_clean not in empty_indicators:
                        meaningful_values.append(val_clean)

                # Náº¿u khÃ´ng cÃ³ giÃ¡ trá»‹ cÃ³ nghÄ©a nÃ o, loáº¡i bá» cá»™t
                if len(meaningful_values) == 0:
                    columns_to_remove.append(col)

        # Loáº¡i bá» cÃ¡c cá»™t trá»‘ng
        if columns_to_remove:
            df_clean = df_clean.drop(columns=columns_to_remove)
            print(f"ğŸ—‘ï¸ ÄÃ£ loáº¡i bá» {len(columns_to_remove)} cá»™t khÃ´ng cÃ³ dá»¯ liá»‡u: {columns_to_remove}")

        return df_clean

    def export_raw_data(self, df, file_path):
        """Xuáº¥t dá»¯ liá»‡u nguyÃªn báº£n khÃ´ng qua validation"""
        if df.empty:
            raise ValueError("KhÃ´ng cÃ³ dá»¯ liá»‡u Ä‘á»ƒ xuáº¥t")

        try:
            # Táº¡o tÃªn sheet vá»›i timestamp
            sheet_name = f"Raw_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            # Loáº¡i bá» cÃ¡c cá»™t khÃ´ng cÃ³ dá»¯ liá»‡u
            df_export = self._remove_empty_columns(df)

            # Chá»‰ Ä‘áº£m báº£o NaN Ä‘Æ°á»£c hiá»ƒn thá»‹ nhÆ° Ã´ trá»‘ng
            df_export = df_export.fillna('')

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Xuáº¥t dá»¯ liá»‡u nguyÃªn báº£n
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)

                # Láº¥y worksheet Ä‘á»ƒ Ä‘á»‹nh dáº¡ng cÆ¡ báº£n
                worksheet = writer.sheets[sheet_name]

                # Chá»‰ Ã¡p dá»¥ng Ä‘á»‹nh dáº¡ng cÆ¡ báº£n (header vÃ  border)
                self._format_raw_worksheet(worksheet, df_export)

            return True, f"ÄÃ£ xuáº¥t dá»¯ liá»‡u nguyÃªn báº£n {len(df)} sinh viÃªn"

        except Exception as e:
            return False, f"Lá»—i xuáº¥t dá»¯ liá»‡u nguyÃªn báº£n: {str(e)}"

    def _format_raw_worksheet(self, worksheet, df):
        """Äá»‹nh dáº¡ng cÆ¡ báº£n cho worksheet dá»¯ liá»‡u nguyÃªn báº£n"""
        # Chá»‰ Ä‘á»‹nh dáº¡ng header
        self._format_header(worksheet)

        # Tá»± Ä‘á»™ng Ä‘iá»u chá»‰nh Ä‘á»™ rá»™ng cá»™t
        self._auto_adjust_columns(worksheet)

        # ThÃªm border cÆ¡ báº£n
        self._add_borders(worksheet, df)
    
    def _format_worksheet(self, worksheet, df):
        """Äá»‹nh dáº¡ng worksheet Excel"""
        # Äá»‹nh dáº¡ng header
        self._format_header(worksheet)
        
        # Tá»± Ä‘á»™ng Ä‘iá»u chá»‰nh Ä‘á»™ rá»™ng cá»™t
        self._auto_adjust_columns(worksheet)
        
        # Äá»‹nh dáº¡ng dá»¯ liá»‡u
        self._format_data_cells(worksheet, df)
        
        # ThÃªm border
        self._add_borders(worksheet, df)
    
    def _format_header(self, worksheet):
        """Äá»‹nh dáº¡ng header"""
        header_font = Font(
            name='Arial',
            size=12,
            bold=True,
            color='FFFFFF'
        )
        
        header_fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        
        # Ãp dá»¥ng Ä‘á»‹nh dáº¡ng cho hÃ ng Ä‘áº§u tiÃªn
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _auto_adjust_columns(self, worksheet):
        """Tá»± Ä‘á»™ng Ä‘iá»u chá»‰nh Ä‘á»™ rá»™ng cá»™t"""
        column_widths = {
            'A': 8,   # STT
            'B': 15,  # Lá»›p
            'C': 15,  # MSV
            'D': 20,  # Há» vÃ  Ä‘á»‡m
            'E': 12,  # TÃªn
            'F': 8,   # CC
            'G': 8    # KT1
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
    
    def _format_data_cells(self, worksheet, df):
        """Äá»‹nh dáº¡ng cÃ¡c Ã´ dá»¯ liá»‡u"""
        data_font = Font(name='Arial', size=11)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Äá»‹nh dáº¡ng tá»«ng cá»™t
        for row_idx in range(2, len(df) + 2):  # Báº¯t Ä‘áº§u tá»« hÃ ng 2
            # STT - center
            worksheet[f'A{row_idx}'].font = data_font
            worksheet[f'A{row_idx}'].alignment = center_alignment
            
            # Lá»›p - center
            worksheet[f'B{row_idx}'].font = data_font
            worksheet[f'B{row_idx}'].alignment = center_alignment
            
            # MSV - center
            worksheet[f'C{row_idx}'].font = data_font
            worksheet[f'C{row_idx}'].alignment = center_alignment
            
            # Há» vÃ  Ä‘á»‡m - left
            worksheet[f'D{row_idx}'].font = data_font
            worksheet[f'D{row_idx}'].alignment = left_alignment
            
            # TÃªn - left
            worksheet[f'E{row_idx}'].font = data_font
            worksheet[f'E{row_idx}'].alignment = left_alignment
            
            # Äiá»ƒm CC - center
            worksheet[f'F{row_idx}'].font = data_font
            worksheet[f'F{row_idx}'].alignment = center_alignment
            
            # Äiá»ƒm KT1 - center
            worksheet[f'G{row_idx}'].font = data_font
            worksheet[f'G{row_idx}'].alignment = center_alignment
    
    def _add_borders(self, worksheet, df):
        """ThÃªm border cho báº£ng"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ãp dá»¥ng border cho táº¥t cáº£ Ã´ cÃ³ dá»¯ liá»‡u
        for row_idx in range(1, len(df) + 2):
            for col_idx in range(1, len(EXCEL_HEADERS) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
    
    def create_summary_sheet(self, writer, df, sheet_name):
        """Táº¡o sheet tÃ³m táº¯t thá»‘ng kÃª"""
        try:
            summary_data = self._generate_summary_data(df)
            
            # Táº¡o DataFrame cho summary
            summary_df = pd.DataFrame(summary_data)
            
            # Xuáº¥t summary sheet
            summary_sheet_name = f"ThongKe_{sheet_name}"
            summary_df.to_excel(writer, sheet_name=summary_sheet_name, index=False)
            
            return True
        except Exception as e:
            print(f"Lá»—i táº¡o summary sheet: {e}")
            return False
    
    def _generate_summary_data(self, df):
        """Táº¡o dá»¯ liá»‡u thá»‘ng kÃª"""
        summary = []
        
        # Thá»‘ng kÃª tá»•ng quan
        summary.append(["Thá»‘ng kÃª", "GiÃ¡ trá»‹"])
        summary.append(["Tá»•ng sá»‘ sinh viÃªn", len(df)])
        summary.append(["", ""])
        
        # Thá»‘ng kÃª theo lá»›p
        if 'Lá»›p' in df.columns:
            class_counts = df['Lá»›p'].value_counts()
            summary.append(["PhÃ¢n bá»‘ theo lá»›p", ""])
            for class_name, count in class_counts.items():
                summary.append([class_name, count])
            summary.append(["", ""])
        
        # Thá»‘ng kÃª Ä‘iá»ƒm
        try:
            if 'CC' in df.columns:
                cc_scores = pd.to_numeric(df['CC'], errors='coerce').dropna()
                if len(cc_scores) > 0:
                    summary.append(["Thá»‘ng kÃª Ä‘iá»ƒm CC", ""])
                    summary.append(["Trung bÃ¬nh", f"{cc_scores.mean():.2f}"])
                    summary.append(["Cao nháº¥t", f"{cc_scores.max():.1f}"])
                    summary.append(["Tháº¥p nháº¥t", f"{cc_scores.min():.1f}"])
                    summary.append(["Äiá»ƒm 10", (cc_scores == 10).sum()])
                    summary.append(["Äiá»ƒm 0", (cc_scores == 0).sum()])
                    summary.append(["", ""])
            
            if 'KT1' in df.columns:
                kt1_scores = pd.to_numeric(df['KT1'], errors='coerce').dropna()
                if len(kt1_scores) > 0:
                    summary.append(["Thá»‘ng kÃª Ä‘iá»ƒm KT1", ""])
                    summary.append(["Trung bÃ¬nh", f"{kt1_scores.mean():.2f}"])
                    summary.append(["Cao nháº¥t", f"{kt1_scores.max():.1f}"])
                    summary.append(["Tháº¥p nháº¥t", f"{kt1_scores.min():.1f}"])
                    summary.append(["Äiá»ƒm 10", (kt1_scores == 10).sum()])
                    summary.append(["Äiá»ƒm 0", (kt1_scores == 0).sum()])
        except Exception as e:
            print(f"Lá»—i tÃ­nh thá»‘ng kÃª Ä‘iá»ƒm: {e}")
        
        return summary
